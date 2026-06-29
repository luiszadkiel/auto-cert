"""
services/server_credentials_service.py
────────────────────────────────────────
Lee las credenciales de servidores mqsi desde un archivo Excel (servers.xlsx).

Estructura del Excel:
  Columna A: JKS_Name   → ej. "Bancaseguro.jks"
  Columna B: Ambiente   → ej. "DEV" | "SQA" | "PROD"
  Columna C: Servidor   → ej. "mqsi-dev-01.bhd.local" (referencia, no se usa en el form)
  Columna D: Usuario    → usuario mqsi para ese servidor
  Columna E: Password   → contraseña mqsi para ese servidor

Lookup: (jks_name, ambiente) → (username, password)

Si no se encuentra una fila exacta, intenta por ambiente solo.
Si tampoco hay coincidencia, retorna ("", "") y loguea un warning.
"""

import os
from dataclasses import dataclass
from typing import Optional

SERVERS_FILE = os.path.join(os.path.dirname(__file__), "..", "servers.xlsx")

# Nombres de columna esperados en la primera fila (case-insensitive)
COL_JKS      = "jks_name"
COL_AMBIENTE = "ambiente"
COL_SERVER   = "servidor"
COL_USER     = "usuario"
COL_PASS     = "password"

REQUIRED_COLS = {COL_JKS, COL_AMBIENTE, COL_USER, COL_PASS}


@dataclass(frozen=True)
class ServerCredential:
    jks_name: str
    ambiente: str
    servidor: str
    username: str
    password: str


class ServerCredentialsService:
    """
    Lee servers.xlsx y proporciona credenciales mqsi por (jks_name, ambiente).
    Crea el archivo de plantilla si no existe.
    """

    def __init__(self, path: str = SERVERS_FILE):
        self._path = os.path.abspath(path)
        self._credentials: list[ServerCredential] = []
        self._loaded = False

    def load(self) -> "ServerCredentialsService":
        """Lee el Excel. Crea la plantilla si no existe. Retorna self para chaining."""
        if not os.path.exists(self._path):
            print(f"  [Credentials] WARN servers.xlsx no encontrado -> creando plantilla en:")
            print(f"    {self._path}")
            _create_template(self._path)
            self._loaded = True
            return self

        try:
            import openpyxl  # type: ignore[import-untyped]
            wb = openpyxl.load_workbook(self._path, read_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                print("  [Credentials] WARN servers.xlsx vacío.")
                return self

            # Leer cabeceras (primera fila)
            headers = [str(h).lower().strip() if h else "" for h in rows[0]]
            missing = REQUIRED_COLS - set(headers)
            if missing:
                print(f"  [Credentials] WARN Columnas faltantes en servers.xlsx: {missing}")
                return self

            idx = {col: headers.index(col) for col in headers if col in REQUIRED_COLS | {COL_SERVER}}

            for row in rows[1:]:
                jks  = str(row[idx[COL_JKS]]).strip()      if row[idx[COL_JKS]]  else ""
                amb  = str(row[idx[COL_AMBIENTE]]).strip()  if row[idx[COL_AMBIENTE]] else ""
                serv = str(row[idx[COL_SERVER]]).strip()    if COL_SERVER in idx and row[idx[COL_SERVER]] else ""
                user = str(row[idx[COL_USER]]).strip()      if row[idx[COL_USER]] else ""
                pwd  = str(row[idx[COL_PASS]]).strip()      if row[idx[COL_PASS]] else ""

                if jks and amb:
                    self._credentials.append(
                        ServerCredential(
                            jks_name=jks,
                            ambiente=amb,
                            servidor=serv,
                            username=user,
                            password=pwd,
                        )
                    )

            wb.close()
            print(f"  [Credentials] OK {len(self._credentials)} servidores cargados desde servers.xlsx")

        except Exception as exc:
            print(f"  [Credentials] ERROR Error al leer servers.xlsx: {exc}")

        self._loaded = True
        return self

    def get(self, jks_name: str, ambiente: str) -> tuple[str, str]:
        """
        Retorna (username, password) para el par (jks_name, ambiente).

        Prioridad de búsqueda:
          1. Coincidencia exacta jks_name + ambiente
          2. Coincidencia solo por ambiente (wildcard para todos los JKS del ambiente)
          3. ("", "") si no hay nada — el uploader usará las credenciales del .env como fallback
        """
        if not self._loaded:
            self.load()

        jks_norm = jks_name.strip().lower()
        amb_norm = ambiente.strip().lower()

        # 1. Exacto
        for cred in self._credentials:
            if cred.jks_name.lower() == jks_norm and cred.ambiente.lower() == amb_norm:
                return cred.username, cred.password

        # 2. Solo ambiente (JKS_Name = "*" o vacío en el Excel)
        for cred in self._credentials:
            if cred.ambiente.lower() == amb_norm and cred.jks_name in ("*", ""):
                return cred.username, cred.password

        # 3. No encontrado
        print(f"  [Credentials] WARN Sin credenciales para '{jks_name}' [{ambiente}] — usando fallback del .env")
        return "", ""

    def all_credentials(self) -> list[ServerCredential]:
        if not self._loaded:
            self.load()
        return list(self._credentials)


def _create_template(path: str) -> None:
    """Crea un servers.xlsx de plantilla con ejemplos y formato."""
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore[import-untyped]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Servidores mqsi"

    # Cabeceras
    headers = ["JKS_Name", "Ambiente", "Servidor", "Usuario", "Password"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas de ejemplo
    examples = [
        ["Bancaseguro.jks", "DEV",  "mqsi-dev-01.bhd.local",  "admin_dev",  "pass_dev"],
        ["Bancaseguro.jks", "SQA",  "mqsi-sqa-01.bhd.local",  "admin_sqa",  "pass_sqa"],
        ["ad.jks",          "SQA",  "mqsi-sqa-01.bhd.local",  "admin_sqa",  "pass_sqa"],
        ["brokerKeystore.jks", "SQA", "mqsi-sqa-02.bhd.local", "admin_sqa2", "pass_sqa2"],
        # Wildcard: todas las JKS del ambiente PROD usan la misma credencial
        ["*",               "PROD", "mqsi-prod-01.bhd.local", "admin_prod", "pass_prod"],
    ]

    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Ancho de columnas
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Hoja de instrucciones
    ws_info = wb.create_sheet("Instrucciones")
    instructions = [
        ["Campo", "Descripción"],
        ["JKS_Name", "Nombre exacto del archivo JKS (ej: Bancaseguro.jks). Usa * para aplicar a todos los JKS del ambiente."],
        ["Ambiente", "Ambiente del servidor: DEV, SQA, PROD, etc."],
        ["Servidor",  "Nombre del host mqsi (solo referencia, no se usa en el portal)."],
        ["Usuario",   "Usuario de acceso al servidor mqsi."],
        ["Password",  "Contraseña del servidor mqsi."],
    ]
    for row_data in instructions:
        ws_info.append(row_data)
    ws_info.column_dimensions["A"].width = 15
    ws_info.column_dimensions["B"].width = 80

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"  [Credentials] OK Plantilla creada: {path}")
    print("  -> Rellena con las credenciales reales y vuelve a ejecutar.")
