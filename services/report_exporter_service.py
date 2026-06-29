"""
services/report_exporter_service.py
─────────────────────────────────────
Genera el Excel (.xlsx) final del inventario de certificados usando openpyxl.

Hojas:
  1. "Inventario" — Las 10 columnas del formato original con formato condicional por estado
  2. "Resumen"    — Totales por estado, clusters, fecha de ejecución

Colores por estado:
  🔴 #FFC7CE  Vencido
  🟢 #C6EFCE  Activo
  🟡 #FFEB9C  Renovado
  🔵 #BDD7EE  Nuevo
  ⚪ #D9D9D9  No encontrado
"""

import os
from datetime import datetime

from models.inventory_cert import (
    InventoryCert,
    EXCEL_COLUMNS,
    STATUS_ACTIVO,
    STATUS_VENCIDO,
    STATUS_RENOVADO,
    STATUS_NO_ENCONTRADO,
    STATUS_NUEVO,
    ALL_STATUSES,
)


# ─── Colores por estado ──────────────────────────────────────────────────────

STATUS_COLORS = {
    STATUS_ACTIVO:         "C6EFCE",   # Verde claro
    STATUS_VENCIDO:        "FFC7CE",   # Rojo claro
    STATUS_RENOVADO:       "FFEB9C",   # Amarillo
    STATUS_NUEVO:          "BDD7EE",   # Azul claro
    STATUS_NO_ENCONTRADO:  "D9D9D9",   # Gris
}

STATUS_FONT_COLORS = {
    STATUS_ACTIVO:         "006100",   # Verde oscuro
    STATUS_VENCIDO:        "9C0006",   # Rojo oscuro
    STATUS_RENOVADO:       "9C5700",   # Naranja oscuro
    STATUS_NUEVO:          "003399",   # Azul oscuro
    STATUS_NO_ENCONTRADO:  "333333",   # Gris oscuro
}

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"


class ReportExporterService:
    """Genera el Excel final del inventario de certificados."""

    def export(
        self,
        certs: list[InventoryCert],
        output_dir: str,
        filename_prefix: str = "Inventario_Certificados_No_Prod",
    ) -> str:
        """
        Genera el XLSX con hoja Inventario + hoja Resumen.

        Args:
            certs: Lista de InventoryCert (resultado de la comparación triple)
            output_dir: Directorio de salida
            filename_prefix: Prefijo del nombre del archivo

        Returns:
            Ruta absoluta del archivo generado
        """
        try:
            import openpyxl  # type: ignore[import-untyped]
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
        except ImportError:
            print("  [Report] ERROR openpyxl no instalado. Usa: pip install openpyxl")
            return ""

        stamp = datetime.now().strftime("%Y-%m-%d")
        filename = f"{filename_prefix}_{stamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
        os.makedirs(output_dir, exist_ok=True)

        wb = openpyxl.Workbook()

        # ── Hoja 1: Inventario ────────────────────────────────────────────
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Inventario")
        ws.title = "Inventario"
        self._write_inventory_sheet(ws, certs, Font, PatternFill, Alignment, Border, Side)

        # ── Hoja 2: Resumen ───────────────────────────────────────────────
        ws_summary = wb.create_sheet("Resumen")
        self._write_summary_sheet(ws_summary, certs, Font, PatternFill, Alignment, Border, Side)

        # ── Guardar ───────────────────────────────────────────────────────
        wb.save(output_path)
        abs_path = os.path.abspath(output_path)
        print(f"  [Report] OK Excel generado: {abs_path}")
        return abs_path

    def _write_inventory_sheet(self, ws, certs, Font, PatternFill, Alignment, Border, Side):
        """Escribe la hoja principal del inventario."""

        # Header
        header_fill = PatternFill("solid", fgColor=HEADER_BG)
        header_font = Font(bold=True, color=HEADER_FG, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, cert in enumerate(certs, 2):
            values = cert.to_excel_row()
            status = cert.status

            # Color de fondo según estado
            bg_color = STATUS_COLORS.get(status, "FFFFFF")
            font_color = STATUS_FONT_COLORS.get(status, "000000")
            row_fill = PatternFill("solid", fgColor=bg_color)
            row_font = Font(color=font_color, size=10)

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.font = row_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-width
        for col_idx, header in enumerate(EXCEL_COLUMNS, 1):
            max_len = len(header)
            for row_idx in range(2, min(len(certs) + 2, 52)):  # Sample first 50 rows
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = min(max_len + 3, 60)

        # Ajustar anchos específicos
        from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
        col_widths = {
            1: 45,   # Cluster
            2: 10,   # Ambiente
            3: 20,   # Namespace
            4: 30,   # Nombre del Secreto
            5: 55,   # Nombre del certificado (CN largo)
            6: 22,   # Fecha de Creación
            7: 22,   # Fecha de Vencimiento
            8: 15,   # Tipo de Secreto
            9: 16,   # Estado
            10: 15,  # Responsable
        }
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze panes (congelar header)
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{len(certs) + 1}"

    def _write_summary_sheet(self, ws, certs, Font, PatternFill, Alignment, Border, Side):
        """Escribe la hoja de resumen con totales y métricas."""

        header_fill = PatternFill("solid", fgColor=HEADER_BG)
        header_font = Font(bold=True, color=HEADER_FG, size=11)
        value_font = Font(size=11)
        bold_font = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Título
        ws.cell(row=1, column=1, value="Métrica").fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).border = thin_border
        ws.cell(row=1, column=2, value="Valor").fill = header_fill
        ws.cell(row=1, column=2).font = header_font
        ws.cell(row=1, column=2).border = thin_border

        # Conteos
        counters = {s: 0 for s in ALL_STATUSES}
        clusters = set()
        namespaces = set()
        secrets = set()

        for cert in certs:
            counters[cert.status] = counters.get(cert.status, 0) + 1
            if cert.cluster:
                clusters.add(cert.cluster)
            if cert.namespace:
                namespaces.add(cert.namespace)
            if cert.secret_name:
                secrets.add(f"{cert.namespace}/{cert.secret_name}")

        metrics = [
            ("Total certificados", len(certs)),
            ("", ""),
            ("Activos", counters[STATUS_ACTIVO]),
            ("Vencidos", counters[STATUS_VENCIDO]),
            ("Renovados (sincronizados al portal)", counters[STATUS_RENOVADO]),
            ("Nuevos (solo en cluster)", counters[STATUS_NUEVO]),
            ("No encontrados (solo en Excel/Portal)", counters[STATUS_NO_ENCONTRADO]),
            ("", ""),
            ("Clusters procesados", len(clusters)),
            ("Namespaces únicos", len(namespaces)),
            ("Secrets únicos", len(secrets)),
            ("", ""),
            ("Cluster(s)", ", ".join(sorted(clusters)) if clusters else "N/A"),
            ("Fecha de ejecución", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for row_idx, (metric, value) in enumerate(metrics, 2):
            cell_m = ws.cell(row=row_idx, column=1, value=metric)
            cell_v = ws.cell(row=row_idx, column=2, value=value)
            cell_m.font = bold_font if metric else value_font
            cell_v.font = value_font
            cell_m.border = thin_border
            cell_v.border = thin_border

            # Colorear filas de estado
            if metric in STATUS_COLORS:
                color = STATUS_COLORS[metric]
                cell_m.fill = PatternFill("solid", fgColor=color)
                cell_v.fill = PatternFill("solid", fgColor=color)

        # Ancho de columnas
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 55
